import streamlit as st
from pymongo import MongoClient
from datetime import datetime, timedelta, timezone, date, time
from bson import ObjectId
import pandas as pd, os, copy
from apscheduler.schedulers.background import BackgroundScheduler
from streamlit_option_menu import option_menu
from streamlit_ws_localstorage import injectWebsocketCode, getOrCreateUID
from io import BytesIO
import base64
import bcrypt
import smtplib
import os
from email.mime.text import MIMEText
import secrets
import openpyxl
import certifi
import plotly.express as px
import plotly.graph_objects as go
from pymongo import MongoClient
from dotenv import load_dotenv
load_dotenv()


MONGO_URI = os.getenv("MONGO_URI")

client = MongoClient(
    MONGO_URI,
    tls=True,
    tlsAllowInvalidCertificates=False,  # ❌ Turn this off for certifi
    tlsCAFile=certifi.where()           # ✅ Use certifi's CA bundle
)

db = client[os.getenv("MONGO_DB")]
collection = db["user_finance"]
users_col = db["users"]
reset_tokens_col = db["reset_tokens"]
data_entries_col = db["data_entries"]

LOG_FILE = os.path.join(os.getcwd(), 'update_log.xlsx')

if not users_col.find_one({"username": "admin"}):
    users_col.insert_one({"username": "admin", "password": bcrypt.hashpw("admin123".encode(), bcrypt.gensalt())})

# Safe conversion
def safe_float(v):
    return float(v) if isinstance(v, (int, float)) or (v and str(v).replace('.', '', 1).replace(',', '').isdigit()) else 0.0

# Email Configuration
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = "thanusrisrini@gmail.com"
SENDER_PASSWORD = "qbzw volu towg tthe"  # Gmail App Password

def send_reset_email(recipient_email, token):
    reset_url = f"http://localhost:8501/?reset_token={token}"
    msg = MIMEText(f"Click the link below to reset your password:\n{reset_url}")
    msg['Subject'] = "Password Reset - Finance Tracker"
    msg['From'] = SENDER_EMAIL
    msg['To'] = recipient_email
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.sendmail(SENDER_EMAIL, recipient_email, msg.as_string())
        return True
    except Exception as e:
        print("Error sending email:", e)
        return False

# ---------------- Dashboard Styling ----------------
def load_dashboard_theme():
    st.markdown(f"""
    <style>
    :root {{
        --primary: #1a103d;  /* Deep royal purple */
        --secondary: #0d0826;  /* Darker purple */
        --accent: #d4af37;  /* Gold accent */
        --success: #3ab795;  /* Emerald green */
        --warning: #e58c3a;  /* Amber */
        --text: #ffffff;  /* White text */
        --highlight: #f8f3e0;  /* Cream highlight */
        --sidebar-width: 300px;
        --bg-gradient: linear-gradient(135deg, #1a103d 0%, #0d0826 100%);
        --card-bg: rgba(26, 16, 61, 0.95);  /* Semi-transparent primary */
        --glass-effect: linear-gradient(135deg, rgba(255, 255, 255, 0.05), rgba(255, 255, 255, 0.01));
    }}

    /* Main app with background image */
    .stApp {{
        background-attachment: fixed;
        font-family: 'Playfair Display', serif;
        color: var(--text);
    }}
    
    /* ============ GRAND SIDEBAR ENHANCEMENTS ============ */
    /* Main sidebar container */
    .st-emotion-cache-6qob1r {{
        background: linear-gradient(160deg, var(--secondary), #140b33) !important;
        border-right: 1px solid var(--accent) !important;
        box-shadow: 8px 0 30px rgba(0, 0, 0, 0.5) !important;
        min-width: var(--sidebar-width) !important;
        max-width: var(--sidebar-width) !important;
        backdrop-filter: blur(16px) !important;
        padding-top: 0 !important;
    }}
    
    /* Sidebar header with gold embellishment */
    .sidebar-header {{
        padding: 2rem 1rem 1.5rem;
        margin-bottom: 1.5rem;
        border-bottom: 1px solid rgba(212, 175, 55, 0.3);
        text-align: center;
        position: relative;
        overflow: hidden;
    }}
    
    .sidebar-header:before {{
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 4px;
        background: linear-gradient(90deg, 
                        transparent 0%, 
                        var(--accent) 50%, 
                        transparent 100%);
    }}
    
    .sidebar-header h2 {{
        color: var(--highlight);
        font-family: 'Playfair Display', serif;
        font-size: 1.8rem;
        margin: 0;
        letter-spacing: 1px;
        position: relative;
        display: inline-block;
        text-shadow: 0 2px 4px rgba(0,0,0,0.5);
    }}
    
    .sidebar-header h2:after {{
        content: '';
        position: absolute;
        bottom: -12px;
        left: 25%;
        width: 50%;
        height: 2px;
        background: linear-gradient(90deg, transparent, var(--accent), transparent);
    }}
    
    /* Sidebar user profile with gold frame */
    .sidebar-profile {{
        display: flex;
        align-items: center;
        padding: 1.5rem;
        margin: 0 1rem 2rem;
        background: rgba(212, 175, 55, 0.1);
        border-radius: 16px;
        border-left: 4px solid var(--accent);
        position: relative;
        overflow: hidden;
        box-shadow: 0 8px 24px rgba(0, 0, 0, 0.3);
    }}
    
    .sidebar-profile:before {{
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 2px;
        background: linear-gradient(90deg, transparent, var(--accent), transparent);
    }}
    
    .sidebar-profile img {{
        width: 60px;
        height: 60px;
        border-radius: 50%;
        border: 3px solid var(--accent);
        margin-right: 1.5rem;
        object-fit: cover;
        box-shadow: 0 4px 12px rgba(212, 175, 55, 0.4);
    }}
    
    .sidebar-profile-info h4 {{
        margin: 0;
        color: var(--highlight);
        font-family: 'Montserrat', sans-serif;
        font-size: 1.1rem;
        letter-spacing: 0.5px;
    }}
    
    .sidebar-profile-info p {{
        margin: 0.25rem 0 0;
        color: rgba(255,255,255,0.8);
        font-size: 0.85rem;
        font-family: 'Montserrat', sans-serif;
    }}
    
    /* Sidebar menu items with gold accents */
    .st-emotion-cache-1wivap2 {{
        padding: 0.5rem 1.5rem !important;
    }}
    
    .st-emotion-cache-1wivap2 a {{
        color: var(--highlight) !important;
        transition: all 0.3s ease !important;
        font-family: 'Montserrat', sans-serif;
        padding: 14px 18px !important;
        border-left: 4px solid transparent;
        border-radius: 12px;
        margin: 6px 0 !important;
        display: flex !important;
        align-items: center !important;
        font-size: 1rem;
        letter-spacing: 0.5px;
        background: var(--glass-effect);
        backdrop-filter: blur(5px);
    }}
    
    .st-emotion-cache-1wivap2 a:hover {{
        color: var(--accent) !important;
        transform: translateX(10px) !important;
        border-left: 4px solid var(--accent) !important;
        background: rgba(212, 175, 55, 0.15) !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.2);
    }}
    
    .st-emotion-cache-1wivap2 a svg {{
        margin-right: 14px !important;
        color: var(--accent) !important;
        font-size: 1.2rem;
    }}
    
    /* Active menu item */
    .st-emotion-cache-1wivap2 a.active {{
        background: rgba(212, 175, 55, 0.25) !important;
        border-left: 4px solid var(--accent) !important;
        color: var(--accent) !important;
        font-weight: 600;
    }}
    
    /* Sidebar footer with subtle pattern */
    .sidebar-footer {{
        position: absolute;
        bottom: 0;
        left: 0;
        right: 0;
        padding: 1.5rem;
        text-align: center;
        border-top: 1px solid rgba(212, 175, 55, 0.2);
        font-size: 0.8rem;
        color: rgba(255,255,255,0.6);
        background: url('https://www.transparenttextures.com/patterns/subtle-white-feathers.png');
        background-size: 200px;
        background-blend-mode: overlay;
        opacity: 0.8;
    }}
    
    /* ============ LUXURY MAIN CONTENT ============ */
    /* Main content area adjustment */
    .main .block-container {{
        padding-left: calc(var(--sidebar-width) + 3rem) !important;
        padding-right: 4rem !important;
        padding-top: 3rem !important;
    }}
    
    /* Dashboard header with logo and crown */
    .dashboard-header-container {{
        display: flex;
        align-items: center;
        margin-bottom: 3rem;
        position: relative;
    }}
    
    .dashboard-header-container:before {{
        content: '';
        position: absolute;
        bottom: -10px;
        left: 0;
        right: 0;
        height: 2px;
        background: linear-gradient(90deg, transparent, var(--accent), transparent);
    }}
    
    .dashboard-logo {{
        width: 70px;
        height: 70px;
        margin-right: 2rem;
        background: linear-gradient(145deg, var(--accent), #e8c252);
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        box-shadow: 0 8px 24px rgba(212, 175, 55, 0.5);
        position: relative;
        overflow: hidden;
    }}
    
    .dashboard-logo:before {{
        content: '';
        position: absolute;
        top: -10px;
        left: -10px;
        right: -10px;
        bottom: -10px;
        border: 2px solid rgba(212, 175, 55, 0.3);
        border-radius: 50%;
        pointer-events: none;
    }}
    
    .dashboard-logo span {{
        font-family: 'Playfair Display', serif;
        font-size: 2rem;
        font-weight: bold;
        color: var(--secondary);
        text-shadow: 0 2px 4px rgba(0,0,0,0.2);
    }}
    
    /* Status indicators with gemstone effect */
    .status-indicator {{
        display: inline-flex;
        align-items: center;
        padding: 8px 16px;
        border-radius: 24px;
        font-size: 0.85rem;
        font-weight: 600;
        margin-left: 1.5rem;
        font-family: 'Montserrat', sans-serif;
        letter-spacing: 0.5px;
        text-transform: uppercase;
        box-shadow: 0 4px 12px rgba(0,0,0,0.2);
        position: relative;
        overflow: hidden;
    }}
    
    .status-indicator:before {{
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 2px;
        background: rgba(255,255,255,0.3);
    }}
    
    .status-active {{
        background: linear-gradient(145deg, rgba(58, 183, 149, 0.3), rgba(58, 183, 149, 0.15));
        color: var(--highlight);
        border: 1px solid var(--success);
    }}
    
    .status-pending {{
        background: linear-gradient(145deg, rgba(229, 140, 58, 0.3), rgba(229, 140, 58, 0.15));
        color: var(--highlight);
        border: 1px solid var(--warning);
    }}
    
    /* Floating action button with glow effect */
    .floating-action-btn {{
        position: fixed;
        bottom: 3rem;
        right: 3rem;
        width: 70px;
        height: 70px;
        border-radius: 50%;
        background: linear-gradient(145deg, var(--accent), #e8c252);
        color: var(--secondary);
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1.8rem;
        box-shadow: 0 8px 30px rgba(212, 175, 55, 0.6);
        cursor: pointer;
        transition: all 0.3s ease;
        z-index: 100;
        border: none;
        animation: pulse 2s infinite;
    }}
    
    @keyframes pulse {{
        0% {{ box-shadow: 0 0 0 0 rgba(212, 175, 55, 0.7); }}
        70% {{ box-shadow: 0 0 0 15px rgba(212, 175, 55, 0); }}
        100% {{ box-shadow: 0 0 0 0 rgba(212, 175, 55, 0); }}
    }}
    
    .floating-action-btn:hover {{
        transform: translateY(-5px) scale(1.1);
        box-shadow: 0 12px 35px rgba(212, 175, 55, 0.8);
        animation: none;
    }}
    
    /* Notification badge with gem cut */
    .notification-badge {{
        position: absolute;
        top: -8px;
        right: -8px;
        background: var(--warning);
        color: white;
        border-radius: 50%;
        width: 24px;
        height: 24px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 0.8rem;
        font-weight: bold;
        box-shadow: 0 2px 8px rgba(0,0,0,0.3);
        border: 2px solid rgba(255,255,255,0.3);
    }}
    
    /* ============ LUXURY CARDS ============ */
    /* Premium cards with gold embellishments */
    .dashboard-card {{
        background: var(--card-bg) !important;
        backdrop-filter: blur(16px);
        -webkit-backdrop-filter: blur(16px);
        border-radius: 20px;
        border: 1px solid var(--accent);
        box-shadow: 0 16px 50px 0 rgba(0, 0, 0, 0.5);
        padding: 2.5rem;
        margin-bottom: 3rem;
        transition: all 0.3s ease;
        position: relative;
        overflow: hidden;
    }}
    
    .dashboard-card:before {{
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 5px;
        background: linear-gradient(90deg, 
                        var(--accent) 0%, 
                        rgba(212, 175, 55, 0.6) 50%, 
                        var(--accent) 100%);
    }}
    
    .dashboard-card:after {{
        content: '';
        position: absolute;
        top: 5px;
        left: 5px;
        right: 5px;
        bottom: 5px;
        border: 1px solid rgba(212, 175, 55, 0.2);
        border-radius: 16px;
        pointer-events: none;
    }}
    
    .dashboard-card:hover {{
        transform: translateY(-10px);
        box-shadow: 0 20px 60px 0 rgba(0, 0, 0, 0.6);
        border: 1px solid var(--highlight);
    }}
    
    .card-header {{
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 2rem;
        padding-bottom: 1.5rem;
        border-bottom: 1px solid rgba(212, 175, 55, 0.3);
    }}
    
    .card-header h3 {{
        color: var(--highlight);
        font-family: 'Playfair Display', serif;
        font-size: 1.6rem;
        margin: 0;
        letter-spacing: 0.5px;
        text-shadow: 0 2px 4px rgba(0,0,0,0.3);
    }}
    
    .card-actions {{
        display: flex;
        gap: 12px;
    }}
    
    .card-action-btn {{
        background: rgba(212, 175, 55, 0.15) !important;
        border: 1px solid var(--accent) !important;
        color: var(--accent) !important;
        border-radius: 12px !important;
        padding: 8px 16px !important;
        font-size: 0.9rem !important;
        transition: all 0.3s ease !important;
        font-family: 'Montserrat', sans-serif;
        letter-spacing: 0.5px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
    }}
    
    .card-action-btn:hover {{
        background: rgba(212, 175, 55, 0.3) !important;
        transform: translateY(-2px);
    }}
    
    /* ============ METRIC CARDS ============ */
    /* Jewel-like metric cards */
    
    .metric-card {{
        background: linear-gradient(145deg, rgba(26, 16, 61, 0.9), rgba(13, 8, 38, 0.9)) !important;
        border-radius: 16px;
        padding: 1.5rem 0.8rem;  /* Reduced horizontal padding */
        text-align: center;
        border: 1px solid var(--accent);
        transition: all 0.3s ease;
        box-shadow: 0 12px 32px rgba(0, 0, 0, 0.4);
        position: relative;
        overflow: hidden;  /* Prevent overflow */
        min-height: 140px;
        display: flex;
        flex-direction: column;
        justify-content: center;
        width: 100%;  /* Ensure full width usage */
        box-sizing: border-box;
    }}
    
    .metric-card:before {{
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 4px;
        background: linear-gradient(90deg, transparent, var(--accent), transparent);
    }}
    
    .metric-label {{
        font-size: 0.9rem;  /* Slightly smaller */
        color: rgba(255,255,255,0.9);
        letter-spacing: 0.5px;
        text-transform: uppercase;
        font-family: 'Montserrat', sans-serif;
        margin-bottom: 0.5rem;
        line-height: 1.2;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;  /* Add ellipsis for long labels */
    }}
    
    .metric-value {{
        font-size: 1.8rem;  /* Reduced font size */
        font-weight: 700;
        color: var(--highlight);
        margin: 0.25rem 0;
        text-shadow: 0 2px 4px rgba(0,0,0,0.3);
        font-family: 'Playfair Display', serif;
        background: linear-gradient(to right, var(--highlight), var(--accent));
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        word-wrap: break-word;  /* Break long words */
        overflow-wrap: break-word;  /* Better word breaking */
        line-height: 1.1;
        max-width: 100%;
        overflow: hidden;
        display: block;
        hyphens: auto;  /* Auto hyphenation */
    }}
    
    /* Responsive adjustments */
    @media (max-width: 1400px) {{
        .metric-value {{
            font-size: 1.6rem;
        }}
    }}
    
    @media (max-width: 1200px) {{
        .metric-value {{
            font-size: 1.4rem;
        }}
        .metric-card {{
            min-height: 120px;
            padding: 1rem 0.6rem;
        }}
        .metric-label {{
            font-size: 0.8rem;
        }}
    }}
    
    @media (max-width: 768px) {{
        .metric-value {{
            font-size: 1.2rem;
        }}
        .metric-card {{
            min-height: 100px;
            padding: 0.8rem 0.4rem;
        }}
        .metric-label {{
            font-size: 0.75rem;
        }}
    }}
    
    /* ============ DATA TABLES ============ */
    /* Elegant data tables */
    .stDataFrame {{
        border-radius: 20px !important;
        background: rgba(26, 16, 61, 0.9) !important;
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        border: 1px solid var(--accent) !important;
        box-shadow: 0 12px 40px rgba(0, 0, 0, 0.4);
        overflow: hidden;
    }}
    
    /* Table headers */
    .stDataFrame thead th {{
        background: linear-gradient(to bottom, var(--secondary), var(--primary)) !important;
        color: var(--highlight) !important;
        font-family: 'Montserrat', sans-serif;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        font-size: 0.9rem;
        border-bottom: 2px solid var(--accent) !important;
    }}
    
    /* Table cells */
    .stDataFrame tbody td {{
        border-bottom: 1px solid rgba(212, 175, 55, 0.2) !important;
        font-family: 'Montserrat', sans-serif;
    }}
    
    /* Hover effect */
    .stDataFrame tbody tr:hover {{
        background: rgba(212, 175, 55, 0.1) !important;
    }}
    
    /* ============ FORM ELEMENTS ============ */
    /* Stylish form elements */
    .stTextInput>div>div>input,
    .stNumberInput>div>div>input,
    .stDateInput>div>div>input,
    .stSelectbox>div>div>select {{
        background: rgba(255, 255, 255, 0.1) !important;
        border-radius: 12px !important;
        padding: 12px 16px !important; /* Adjusted padding */
        color: var(--highlight) !important;
        border: 1px solid var(--accent) !important;
        font-family: 'Montserrat', sans-serif;
        transition: all 0.3s ease !important;
        /* Font size adjustments */
        font-size: 16px !important; /* Base size for inputs */
        font-weight: normal !important; /* Changed from bold */
        width: 100% !important;
        box-sizing: border-box;
    }}

    /* Larger font for text inputs */
    .stTextInput>div>div>input {{
        font-size: 18px !important;
    }}

    /* Slightly smaller for number inputs */
    .stNumberInput>div>div>input {{
        font-size: 16px !important;
    }}

    /* Date inputs */
    .stDateInput>div>div>input {{
        font-size: 16px !important;
    }}

    /* Select dropdowns */
    .stSelectbox>div>div>select {{
        font-size: 16px !important;
    }}

    /* Labels */
    .stTextInput label,
    .stNumberInput label,
    .stDateInput label,
    .stSelectbox label {{
        font-size: 16px !important;
        color: var(--highlight) !important;
        margin-bottom: 8px !important;
        font-weight: 500 !important;
    }}

    /* ============ OTHER ENHANCEMENTS ============ */
    /* Custom scrollbar */
    ::-webkit-scrollbar {{
        width: 12px;
        height: 12px;
    }}
    
    ::-webkit-scrollbar-track {{
        background: rgba(255, 255, 255, 0.05);
        border-radius: 6px;
    }}
    
    ::-webkit-scrollbar-thumb {{
        background: var(--accent);
        border-radius: 6px;
        border: 2px solid rgba(255,255,255,0.1);
    }}
    
    ::-webkit-scrollbar-thumb:hover {{
        background: var(--highlight);
    }}
    
    /* Tooltips */
    .stTooltip {{
        background: var(--secondary) !important;
        border: 1px solid var(--accent) !important;
        color: var(--highlight) !important;
        font-family: 'Montserrat', sans-serif;
        border-radius: 12px !important;
        box-shadow: 0 8px 24px rgba(0,0,0,0.4) !important;
        backdrop-filter: blur(10px);
        padding: 1rem !important;
    }}
    
    /* Loading spinner */
    .stSpinner>div {{
        border: 5px solid rgba(212, 175, 55, 0.2);
        border-top: 5px solid var(--accent);
        border-bottom: 5px solid var(--accent);
        width: 50px !important;
        height: 50px !important;
    }}
    
    /* Toast notifications */
    .stNotification {{
        background: rgba(26, 16, 61, 0.98) !important;
        border: 1px solid var(--accent) !important;
        border-radius: 16px !important;
        backdrop-filter: blur(12px) !important;
        box-shadow: 0 8px 32px rgba(0,0,0,0.4) !important;
    }}
    
    /* Tabs styling */
    .stTabs [data-baseweb="tab-list"] {{
        gap: 0.5rem;
    }}
    
    .stTabs [data-baseweb="tab"] {{
        background: rgba(26, 16, 61, 0.7) !important;
        border-radius: 12px 12px 0 0 !important;
        padding: 1rem 1.5rem !important;
        font-family: 'Montserrat', sans-serif;
        color: var(--text) !important;
        border: 1px solid rgba(212, 175, 55, 0.3) !important;
        border-bottom: none !important;
        transition: all 0.3s ease !important;
    }}
    
    .stTabs [data-baseweb="tab"]:hover {{
        background: rgba(212, 175, 55, 0.2) !important;
        color: var(--highlight) !important;
    }}
    
    .stTabs [aria-selected="true"] {{
        background: linear-gradient(to bottom, var(--primary), var(--secondary)) !important;
        color: var(--highlight) !important;
        border: 1px solid var(--accent) !important;
        border-bottom: none !important;
        font-weight: bold;
    }}
    
    /* Date picker customization */
    .stDateInput>div>div>input {{
        background: rgba(255, 255, 255, 0.1) !important;
    }}
    
    /* Section headers with gold underline */
    .section-header {{
        font-size: 1.6rem;
        font-weight: 600;
        color: var(--highlight);
        margin-bottom: 2rem;
        padding-bottom: 1rem;
        border-bottom: 2px solid var(--accent);
        font-family: 'Playfair Display', serif;
        letter-spacing: 0.5px;
        position: relative;
    }}
    
    .section-header:after {{
        content: '';
        position: absolute;
        bottom: -2px;
        left: 0;
        width: 100px;
        height: 3px;
        background: linear-gradient(90deg, var(--accent), transparent);
    }}
    
    /* Dashboard header with gold gradient */
    .dashboard-header {{
        font-size: 2.4rem;
        font-weight: 700;
        color: var(--highlight);
        margin-bottom: 3rem;
        text-align: center;
        background: linear-gradient(90deg, 
                        rgba(212, 175, 55, 0.2) 0%, 
                        rgba(67, 97, 238, 0.4) 50%, 
                        rgba(212, 175, 55, 0.2) 100%);
        padding: 1.5rem;
        border-radius: 16px;
        border: 1px solid var(--accent);
        text-shadow: 0 4px 8px rgba(0,0,0,0.4);
        letter-spacing: 1px;
        position: relative;
        overflow: hidden;
    }}
    
    .dashboard-header:before {{
        content: '';
        position: absolute;
        bottom: 0;
        left: 25%;
        width: 50%;
        height: 3px;
        background: linear-gradient(90deg, transparent, var(--accent), transparent);
    }}
    </style>
    """, unsafe_allow_html=True)

# Set a luxurious background image
BACKGROUND_URL = "https://images.unsplash.com/photo-1553729459-efe14ef6055d?ixlib=rb-4.0.3&ixid=M3wxMjA3fDB8MHxwaG90by1wYWdlfHx8fGVufDB8fHx8fA%3D%3D&auto=format&fit=crop&w=2070&q=80"

def set_background(url):
    st.markdown(
        f"""
        <style>
        .stApp {{
            background-image: url("{url}");
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
            background-attachment: fixed;
            background-blend-mode: overlay;
            background-color: rgba(15, 10, 35, 0.95);
        }}
        
        /* Add a subtle gold pattern overlay */
        .stApp:after {{
            content: "";
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: url('https://www.transparenttextures.com/patterns/gold-scale.png');
            opacity: 0.08;
            pointer-events: none;
            z-index: -1;
        }}
        
        /* Main content area adjustment */
        .main .block-container {{
            padding-top: 4rem;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )



# Call these functions after your imports
set_background(BACKGROUND_URL)
load_dashboard_theme()


# ---------------- Helper Functions ----------------
def is_admin_user():
    """Check if the current logged-in user is admin"""
    return login_state.get("username") == "admin"

def get_user_filter():
    """Get database filter based on current user permissions"""
    return {"created_by_user": login_state.get("username")}  # All users see only their own data

def create_metric_card(label, value, delta=None, delta_color="normal"):
    """Create a metric card with optional delta indicator"""
    delta_html = ""
    if delta is not None:
        delta_color = "inverse" if delta_color == "inverse" else "normal"
        delta_html = f"""
        <div style="font-size: 0.9rem; color: {'#4CAF50' if delta >= 0 else '#FF5722'};">
            {'+' if delta >= 0 else ''}{delta:,.2f}
        </div>
        """
    
    return f"""
    <div class="metric-card">
        <div class="metric-label">{label}</div>
        <div class="metric-value">{value}</div>
        {delta_html}
    </div>
    """

def create_dashboard_header(title):
    """Create a styled dashboard header"""
    return f"""
    <div class="dashboard-header">
        {title}
    </div>
    """

def create_section_header(title):
    """Create a styled section header"""
    return f"""
    <div class="section-header">
        {title}
    </div>
    """

# ---------------- Persistent login state ----------------
# Persistent login state
@st.cache_resource
def get_login_state():
    return {"logged_in": False, "username": ""}

login_state = get_login_state()

def login():
    st.markdown(create_dashboard_header("🔐 Finance Portal"), unsafe_allow_html=True)
    if login_state["logged_in"]:
        st.success(f"✅ Logged in as **{login_state['username']}**")
        return

    # Registration (top right)
    top_col1, top_col2 = st.columns([4, 1])
    with top_col2:
        with st.popover("Register", use_container_width=True):
            st.markdown("### 🧾 Create New Account")
            with st.form("register_form", clear_on_submit=True):
                reg_user = st.text_input("Username", key="reg_user")
                reg_email = st.text_input("Email", key="reg_email")
                reg_pass = st.text_input("Password", type="password", key="reg_pass")
                reg_btn = st.form_submit_button("Register")
                if reg_btn:
                    if users_col.find_one({"username": reg_user}):
                        st.error("Username already exists!")
                    elif users_col.find_one({"email": reg_email}):
                        st.error("Email already registered!")
                    else:
                        hashed_pw = bcrypt.hashpw(reg_pass.encode(), bcrypt.gensalt())
                        users_col.insert_one({"username": reg_user, "email": reg_email, "password": hashed_pw})
                        st.success("✅ Registration successful! Please log in.")

    # Login form
    # Login form
    with st.container():
        st.markdown("""
        <div class="dashboard-form">
            <h3 style="color: white;">🔑 Login</h3>
        """, unsafe_allow_html=True)
        
    with st.form("login_form", clear_on_submit=False):
        username = st.text_input("Username", key="login_username")
        password = st.text_input("Password", type="password", key="login_password")
        login_col, forgot_col = st.columns([1, 1.5])
        login_btn = login_col.form_submit_button("Login")

        with forgot_col:
            with st.popover("Forgot Password?", use_container_width=True):
                st.markdown("### 🔐 Reset Password")
                email = st.text_input("Enter your registered email", key="forgot_email")
                send_btn = st.form_submit_button("Send Reset Link")
                if send_btn:
                    user = users_col.find_one({"email": email})
                    if user:
                        token = secrets.token_urlsafe(16)
                        reset_tokens_col.insert_one({"email": email, "token": token, "created_at": datetime.utcnow()})
                        if send_reset_email(email, token):
                            st.success("Password reset link sent to your email.")
                        else:
                            st.error("Failed to send email.")
                    else:
                        st.error("Email not registered.")

    if login_btn:
        user = users_col.find_one({"username": username})
        if user and bcrypt.checkpw(password.encode(), user["password"]):
            login_state["logged_in"] = True
            login_state["username"] = username
            st.success("Login successful!")
            st.rerun()
        else:
            st.error("Invalid username or password.")

def logout():
    if st.sidebar.button("Logout", key="logout_btn"):
        login_state["logged_in"] = False
        login_state["username"] = ""
        st.rerun()

# ---------------- Main Auth Flow ----------------
# ---------------- Main Auth Flow ----------------
params = st.query_params
reset_token = params.get("reset_token", None)

# First handle the reset token if present
if reset_token:
    # If logged in but has reset token, clear it and refresh
    if login_state["logged_in"]:
        del params["reset_token"]  # Remove the reset_token parameter
        st.query_params = params  # Update the query params
        st.rerun()
    
    # Process the reset token
    token_doc = reset_tokens_col.find_one({
        "token": reset_token,
        "created_at": {"$gt": datetime.utcnow() - timedelta(hours=24)}
    })
    
    if token_doc:
        # Show password reset form
        st.markdown("## 🔄 Reset Your Password")
        new_pw = st.text_input("New Password", type="password", key="new_pw")
        confirm_pw = st.text_input("Confirm Password", type="password", key="confirm_pw")
        if st.button("Update Password", key="update_pw"):
            if new_pw != confirm_pw:
                st.error("Passwords do not match!")
            else:
                hashed_pw = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt())
                users_col.update_one({"email": token_doc["email"]}, {"$set": {"password": hashed_pw}})
                reset_tokens_col.delete_one({"token": reset_token})
                st.success("✅ Password updated! Please log in.")
                del params["reset_token"]  # Remove the reset_token parameter
                st.query_params = params  # Update the query params
                st.rerun()
        st.stop()
    else:
        # Show invalid token message and stop execution
        st.warning("Invalid or expired reset link.")
        del params["reset_token"]  # Remove the reset_token parameter
        st.query_params = params  # Update the query params
        st.stop()

# Normal app flow continues here
if not login_state["logged_in"]:
    
    login()
    st.stop()

# Only reached when logged in
logout()
st.sidebar.markdown(f"""
    <div style="background: rgba(46, 46, 72, 0.8); 
            padding: 1rem; 
            border-radius: 12px;
            margin-bottom: 1rem;
            text-align: center;">
    👤  <strong>{login_state['username']}</strong><br>
    👨‍💼 User Access
</div>
""", unsafe_allow_html=True)

# Show user access info
st.sidebar.info("👨‍💼 **User**: You can only see and manage your own data")

# ---------------- Scheduler ----------------
def clean_orphaned_log_entries():
    """Remove log entries for users that no longer exist in the database."""
    if not os.path.exists(LOG_FILE):
        return
    
    df = pd.read_excel(LOG_FILE)
    # Get all existing users (admin can see all, regular users see only their own)
    existing_users = {u['name'] for u in collection.find()}
    
    # Filter out entries for non-existent users
    df_cleaned = df[df['name'].isin(existing_users)]
    
    # Only save if there were changes
    if len(df_cleaned) != len(df):
        df_cleaned.to_excel(LOG_FILE, index=False)
        print(f"Cleaned {len(df) - len(df_cleaned)} orphaned log entries")

def log_action(user, action, amount=0.0, comment="", updates=None, time=None):
    df = pd.read_excel(LOG_FILE) if os.path.exists(LOG_FILE) else pd.DataFrame()

    entry = {
        "name": user['name'],
        "action": action,
        "amount_changed": f"{'+₹' if amount > 0 else '-₹'}{abs(amount):,.2f}" if amount != 0 else None,
        "comment": comment,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "last_amount_update": datetime.now().replace(tzinfo=None),
        "time": time or datetime.now().strftime("%Y-%m-%d"),
        "created_by_user": user.get('created_by_user', login_state.get("username"))  # Add user ownership
    }

    # Add only changed fields from updates
    if updates:
        for k in ["remaining_amount", "monthly_interest_rate", "total_interest"]:
            entry[k] = updates.get(k, None)

    df = pd.concat([df, pd.DataFrame([entry])], ignore_index=True)
    df.to_excel(LOG_FILE, index=False)



def apply_monthly_interest(user, until_date=None, from_date=None):
    """Calculate interest chronologically, respecting payment timing and manual updates"""
    if not user:
        return

    now = until_date or datetime.now()
    
    # Get loan start date
    loan_start_date = user.get('loan_start_date')
    if not loan_start_date:
        return
        
    start = loan_start_date
    
    # Convert dates to proper format
    if isinstance(start, str):
        start = datetime.fromisoformat(start)
    if start.tzinfo:
        start = start.replace(tzinfo=None)
    if now.tzinfo:
        now = now.replace(tzinfo=None)

    # Get all transactions for this user from the log, sorted chronologically
    df_log = pd.read_excel(LOG_FILE) if os.path.exists(LOG_FILE) else pd.DataFrame()
    user_logs = df_log[df_log['name'] == user['name']].copy()
    
    if not user_logs.empty:
        # Parse timestamps and sort chronologically
        user_logs['parsed_time'] = pd.to_datetime(user_logs['time'], errors='coerce')
        user_logs = user_logs.sort_values('parsed_time')
    
    # Remove all auto interest entries - we'll recalculate them
    df_log = df_log[~(
        (df_log['name'] == user['name']) &
        (df_log['action'] == 'Interest Auto Update')
    )]
    df_log.to_excel(LOG_FILE, index=False)
    
    # Build timeline of balance changes
    timeline = []
    current_balance = user['principal']  # Start with original principal
    manual_interest_total = 0
    
    # Process each transaction chronologically
    for _, row in user_logs.iterrows():
        action = row['action']
        parsed_time = row['parsed_time']
        
        # Skip auto interest entries and creation entry
        if action in ['Interest Auto Update', 'Created']:
            continue
            
        # Parse amount
        amount = 0.0
        amount_changed = row.get('amount_changed', '')
        if amount_changed and amount_changed != 'None':
            try:
                amount_str = str(amount_changed).replace('+₹', '').replace('-₹', '').replace(',', '')
                if str(amount_changed).startswith('-'):
                    amount = -float(amount_str)
                else:
                    amount = float(amount_str)
            except:
                amount = 0.0
        
        # Add to timeline
        timeline.append({
            'date': parsed_time,
            'action': action,
            'amount': amount,
            'balance_before': current_balance
        })
        
        # Update running balance and manual interest
        if action == 'Payment':
            current_balance += amount  # amount is already signed (negative for payment)
        elif action == 'Interest Update':
            manual_interest_total += amount
        elif action == 'Interest Rate':
            # Rate changes don't affect balance but will affect future calculations
            pass
    
    # Now calculate auto interest month by month, considering timeline
    current_date = start
    running_interest = manual_interest_total
    
    while current_date < now:
        # Find the balance that was active during this month
        month_start = current_date
        try:
            month_end = current_date + pd.DateOffset(months=1)
        except:
            month_end = current_date.replace(month=current_date.month + 1 if current_date.month < 12 else 1, 
                                           year=current_date.year + (1 if current_date.month == 12 else 0))
        
        if hasattr(month_end, 'tzinfo') and month_end.tzinfo:
            month_end = month_end.replace(tzinfo=None)
        
        # Calculate month boundaries for comparison
        if month_end > now:
            effective_month_end = now
        else:
            effective_month_end = month_end
        
        # Find the balance at the start of this month and active rate
        month_balance = user['principal']  # Start with original principal
        active_rate = user.get('monthly_interest_rate', 0) / 100
        
        # Apply all transactions that happened before this month
        for event in timeline:
            if event['date'] < month_start:
                if event['action'] == 'Payment':
                    month_balance += event['amount']
                elif event['action'] == 'Interest Rate':
                    # Find the rate change amount
                    rate_change = event['amount']
                    active_rate = max(0, active_rate + rate_change / 100)
        
        # Find all transactions within this month
        month_transactions = []
        for event in timeline:
            if month_start <= event['date'] < effective_month_end:
                month_transactions.append(event)
        
        # Calculate interest with mid-month payment consideration
        monthly_interest = 0.0
        current_month_balance = month_balance
        period_start = month_start
        
        # Sort month transactions by date
        month_transactions.sort(key=lambda x: x['date'])
        
        # Process each transaction within the month
        for transaction in month_transactions:
            transaction_date = transaction['date']
            
            # Calculate interest for period before this transaction
            if transaction_date > period_start:
                days_in_period = (transaction_date - period_start).days
                daily_rate = active_rate / 30  # Approximate monthly rate to daily
                period_interest = current_month_balance * daily_rate * days_in_period
                monthly_interest += period_interest
            
            # Update balance and rate if transaction affects them
            if transaction['action'] == 'Payment':
                current_month_balance += transaction['amount']  # amount is signed
            elif transaction['action'] == 'Interest Rate':
                rate_change = transaction['amount']
                active_rate = max(0, active_rate + rate_change / 100)
            
            # Update period start for next calculation
            period_start = transaction_date
        
        # Calculate interest for remaining period in month
        if period_start < effective_month_end:
            remaining_days = (effective_month_end - period_start).days
            daily_rate = active_rate / 30  # Approximate monthly rate to daily
            remaining_interest = current_month_balance * daily_rate * remaining_days
            monthly_interest += remaining_interest
        
        # Round to 2 decimal places
        monthly_interest = round(monthly_interest, 2)
        
        # Prepare log entry details
        if month_end > now:
            month_str = now.strftime("%Y-%m-%d")
            comment = f"Interest for {month_start.strftime('%b %Y')} (partial month)"
        else:
            month_str = month_end.strftime("%Y-%m-%d")
            comment = f"Interest for {month_end.strftime('%b %Y')}"
        
        # Add interest entry if it doesn't exist and interest > 0
        if monthly_interest > 0:
            # Check if this entry already exists
            existing_log = pd.read_excel(LOG_FILE) if os.path.exists(LOG_FILE) else pd.DataFrame()
            existing_mask = (
                (existing_log['name'] == user['name']) &
                (existing_log['action'] == 'Interest Auto Update') &
                (pd.to_datetime(existing_log['time'], errors='coerce').dt.date == 
                 pd.to_datetime(month_str).date())
            )
            
            if not existing_mask.any():
                running_interest += monthly_interest
                
                # Log the interest entry
                log_action(
                    user,
                    "Interest Auto Update",
                    monthly_interest,
                    comment,
                    {
                        "total_interest": f"₹{running_interest:,.2f}",
                        "remaining_amount": f"₹{current_month_balance:,.2f}"
                    },
                    time=month_str
                )
        
        # Move to next month
        current_date = month_end
    
    # Update user's total interest in database
    collection.update_one(
        {"_id": user["_id"]}, 
        {"$set": {
            "total_interest": round(running_interest, 2),
            "last_interest_update": now
        }}
    )


# Delete action function

def delete_action(row_index, user_name):
    """Delete a specific action from the log and recalculate chronologically."""
    if not os.path.exists(LOG_FILE):
        return False
    
    df = pd.read_excel(LOG_FILE)
    if row_index >= len(df):
        return False
    
    row = df.iloc[row_index]
    if row['name'] != user_name:
        return False
    
    deleted_action = row.get('action', '')
    
    # Remove the row from log
    df = df.drop(index=row_index).reset_index(drop=True)
    df.to_excel(LOG_FILE, index=False)
    
    # Get the user from database
    user = collection.find_one({"name": user_name})
    if not user:
        return False
    
    # For any deletion, we need to:
    # 1. Reset user to original state (principal, 0 interest, original rate)
    # 2. Apply chronological recalculation which will replay all remaining transactions
    
    # Reset user to original state
    reset_data = {
        "remaining_amount": user["principal"],
        "total_interest": 0.0,
        "monthly_interest_rate": user.get("original_monthly_interest_rate", user["monthly_interest_rate"]),
        "last_interest_update": user.get("loan_start_date", user.get("created_at", datetime.now())),
    }
    collection.update_one({"_id": user["_id"]}, {"$set": reset_data})
    
    # Get the updated user and recalculate everything chronologically
    user_updated = collection.find_one({"_id": user["_id"]})
    apply_monthly_interest(user_updated, until_date=datetime.now())
    
    return True



scheduler = BackgroundScheduler()
scheduler.add_job(lambda: [apply_monthly_interest(u) for u in collection.find()], 'interval', days=30)
scheduler.start()

# ---------------- UI ----------------
with st.sidebar:
    selected = option_menu(
        menu_title=" Menu ",
        options=["Dashboard", "Add User", "Actions", "View", "Data Entry"],
        icons=["speedometer", "plus-circle", "activity", "bar-chart-line", "file-text"],
        default_index=0,
        styles={
            "container": {
                "padding": "0.5rem",
                "background-color": "rgba(46, 46, 72, 0.8)",
                "border-radius": "16px",
                "box-shadow": "0 8px 32px 0 rgba(0, 0, 0, 0.36)"
            },
            "icon": {"color": "#00aaff", "font-size": "20px"},
            "nav-link": {
                "font-size": "16px",
                "text-align": "left",
                "margin": "0.25rem",
                "color": "#f0f0f0",
                "border-radius": "12px",
                "padding": "0.75rem",
            },
            "nav-link-selected": {
                "background": "linear-gradient(145deg, #4e4e7d, #2c2c54)",
                "box-shadow": "0 4px 12px rgba(0, 0, 0, 0.2)"
            },
        }
    )
    
# Dashboard Page
if selected == "Dashboard":
    st.markdown(create_dashboard_header("🗃️ Finance Dashboard"), unsafe_allow_html=True)
    
    # Summary Metrics
    col1, col2, col3, col4 = st.columns(4)
    
    user_filter = get_user_filter()
    total_users = collection.count_documents(user_filter)
    total_principal = sum([u.get('principal', 0) for u in collection.find(user_filter)])
    total_remaining = sum([u.get('remaining_amount', 0) for u in collection.find(user_filter)])
    total_interest = sum([u.get('total_interest', 0) for u in collection.find(user_filter)])
    
    with col1:
        st.markdown(create_metric_card("Total Users", total_users), unsafe_allow_html=True)
    with col2:
        st.markdown(create_metric_card("Total Principal", f"₹{total_principal:,.2f}"), unsafe_allow_html=True)
    with col3:
        st.markdown(create_metric_card("Total Remaining", f"₹{total_remaining:,.2f}"), unsafe_allow_html=True)
    with col4:
        st.markdown(create_metric_card("Total Interest", f"₹{total_interest:,.2f}"), unsafe_allow_html=True)
    
    # Recent Activity
    st.markdown(create_section_header("📈 Recent Activity"), unsafe_allow_html=True)
    
    if os.path.exists(LOG_FILE):
        df_logs = pd.read_excel(LOG_FILE)
        if not df_logs.empty:
            # Filter logs based on user permissions - all users see only their own data
            if 'created_by_user' in df_logs.columns:
                df_logs = df_logs[df_logs['created_by_user'] == login_state.get("username")]
            else:
                # If created_by_user column doesn't exist, show empty dataframe
                df_logs = df_logs.iloc[0:0]  # Empty dataframe with same structure
            recent_activity = df_logs.sort_values(by='timestamp', ascending=False).head(10)
            st.dataframe(recent_activity, use_container_width=True)
        else:
            st.info("No recent activity found.")
    else:
        st.info("No activity log exists yet.")
    
    # User Distribution Chart
    st.markdown(create_section_header("👥 User Distribution"), unsafe_allow_html=True)
    
    user_data = list(collection.find(user_filter, {"name": 1, "principal": 1, "remaining_amount": 1, "total_interest": 1}))
    if user_data:
        df_users = pd.DataFrame(user_data)
        
        tab1, tab2 = st.tabs(["Principal Distribution", "Remaining vs Interest"])
        
        with tab1:
            fig = px.pie(df_users, values='principal', names='name', 
                          title='Principal Amount by User',
                          color_discrete_sequence=px.colors.qualitative.Pastel)
            fig.update_traces(textposition='inside', textinfo='percent+label')
            st.plotly_chart(fig, use_container_width=True)
        
        with tab2:
            fig = go.Figure(data=[
                go.Bar(name='Remaining', x=df_users['name'], y=df_users['remaining_amount'], marker_color='#4CAF50'),
                go.Bar(name='Interest', x=df_users['name'], y=df_users['total_interest'], marker_color='#FF5722')
            ])
            fig.update_layout(barmode='group', title='Remaining Amount vs Interest by User')
            st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("No user data available for visualization.")


# Add Loan
if selected == "Add User":
    st.markdown(create_dashboard_header("👤 Add New Member"), unsafe_allow_html=True)
    with st.form("add_loan"):
        name = st.text_input("Name")
        principal = st.number_input("Principal", min_value=0.0)
        rate = st.number_input("Monthly Interest Rate (%)", min_value=0.0)
        loan_start = st.date_input("Loan Start Date", value=datetime.today())
        
        # Check for duplicate user in DB or log (within current user's scope)
        user_filter = get_user_filter()
        user_filter["name"] = name
        user_exists = collection.find_one(user_filter)
        log_exists = False
        if os.path.exists(LOG_FILE):
            df_log = pd.read_excel(LOG_FILE)
            # Check if they already have this name in their scope
            if 'created_by_user' in df_log.columns:
                log_exists = ((df_log['name'] == name) & (df_log['action'] == 'Created') & (df_log['created_by_user'] == login_state.get("username"))).any()
            else:
                log_exists = False  # If column doesn't exist, no duplicates found
        
        if st.form_submit_button("Add"):
            if not name.strip():
                st.error("Name is required.")
            elif principal <= 0:
                st.error("Principal amount must be greater than 0.")
            elif rate <= 0:
                st.error("Interest rate must be greater than 0.")
            elif user_exists or log_exists:
                st.error(f"User '{name}' already exists in your records! Please delete the user and their log before creating again.")
            else:
                start_dt = datetime.combine(loan_start, datetime.min.time())
                data = {
                    "name": name.strip(),
                    "principal": principal,
                    "remaining_amount": principal,
                    "total_interest": 0.0,
                    "created_at": datetime.now(),
                    "loan_start_date": start_dt,
                    "monthly_interest_rate": rate,
                    "original_monthly_interest_rate": rate,
                    "created_by_user": login_state.get("username")  # Add user ownership
                }
                collection.insert_one(data)
                # Get the newly created user and apply interest
                new_user = collection.find_one({"name": name, "created_by_user": login_state.get("username")})
                if new_user:
                    apply_monthly_interest(new_user)
                log_action(data, "Created", 0.0, "Loan created")
                st.success("Loan added!")
                st.rerun()

# Actions
elif selected == "Actions":
    st.markdown(create_dashboard_header("⚡ User Actions"), unsafe_allow_html=True)
    user_filter = get_user_filter()
    names = [u['name'] for u in collection.find(user_filter, {"name": 1})]
    if not names:
        st.warning("No users found. Please add a user first.")
        st.stop()
    selected = st.selectbox("Select User", names)
    user_filter["name"] = selected
    user = collection.find_one(user_filter)

    if user:
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(create_metric_card("Principal", f"₹{user['principal']:,.2f}"), unsafe_allow_html=True)
        with col2:
            st.markdown(create_metric_card("Remaining", f"₹{user['remaining_amount']:,.2f}"), unsafe_allow_html=True)
        with col3:
            st.markdown(create_metric_card("Interest", f"₹{user['total_interest']:,.2f}"), unsafe_allow_html=True)


        with st.form("action_form"):
            action = st.selectbox("Action", ["Payment", "Interest Update", "Interest Rate"])
            amt = st.number_input("Amount", min_value=0.0)
            custom_date = st.date_input("Date", value=datetime.today())

            col1, col2 = st.columns(2)
            increment_clicked = col1.form_submit_button("➕ Increment")
            decrement_clicked = col2.form_submit_button("➖ Decrement")

            if increment_clicked or decrement_clicked:
                sign = 1 if increment_clicked else -1
                amt_signed = amt * sign
                # Create timezone-naive datetime for consistent handling
                custom_dt = datetime.combine(custom_date, datetime.min.time())

                updates = {}

                if action == "Payment":
                    if amt_signed > user['remaining_amount']:
                        st.error("Exceeds balance")
                        st.stop()

                    old_remaining = user['remaining_amount']
                    new_remaining = old_remaining + amt_signed

                    # Ensure it doesn't go negative
                    if new_remaining < 0:
                        st.error("Resulting remaining amount cannot be negative!")
                        st.stop()

                    old_interest = user['total_interest']

                    # Log the payment first (chronological recalculation depends on log entries)
                    log_action(user, action, amt_signed, action, {
                        "remaining_amount": f"₹{old_remaining:,.2f}",
                        "total_interest": f"₹{old_interest:,.2f}"
                    }, time=custom_dt.strftime("%Y-%m-%d"))
                    
                    # Now recalculate interest chronologically
                    apply_monthly_interest(user, until_date=datetime.now())

                    # Get the final calculated values after chronological recalculation
                    final_user = collection.find_one({"_id": user['_id']})
                    new_interest = final_user['total_interest']
                    
                    # Update the remaining amount in the database based on chronological calculation
                    collection.update_one(
                        {"_id": user["_id"]},
                        {"$set": {
                            "remaining_amount": new_remaining
                        }}
                    )

                    # Update the log entry with final values
                    if os.path.exists(LOG_FILE):
                        df_log = pd.read_excel(LOG_FILE)
                        # Find the most recent entry for this user and action
                        user_mask = (df_log['name'] == user['name']) & (df_log['action'] == action)
                        if user_mask.any():
                            last_idx = df_log[user_mask].index[-1]
                            df_log.at[last_idx, 'remaining_amount'] = f"₹{new_remaining:,.2f}"
                            df_log.at[last_idx, 'total_interest'] = f"₹{new_interest:,.2f}"
                            df_log.to_excel(LOG_FILE, index=False)

                    # Show that the payment has been processed
                    st.success(f"Payment of ₹{abs(amt_signed):,.2f} recorded and interest recalculated chronologically")
                    st.info(f"Balance: ₹{old_remaining:,.2f} → ₹{new_remaining:,.2f} | Interest: ₹{old_interest:,.2f} → ₹{new_interest:,.2f}")
                    st.rerun()

                elif action == "Interest Update":
                    old_interest = user['total_interest']
                    
                    # Log the interest update first (chronological recalculation depends on log entries)
                    log_action(user, action, amt_signed, action, {
                        "remaining_amount": f"₹{user['remaining_amount']:,.2f}",
                        "total_interest": f"₹{old_interest:,.2f}"
                    }, time=custom_dt.strftime("%Y-%m-%d"))
                    
                    # Now recalculate interest chronologically
                    apply_monthly_interest(user, until_date=datetime.now())
                    
                    # Get the final calculated values
                    final_user = collection.find_one({"_id": user['_id']})
                    new_interest = final_user['total_interest']
                    
                    # Update the log entry with final values
                    if os.path.exists(LOG_FILE):
                        df_log = pd.read_excel(LOG_FILE)
                        # Find the most recent entry for this user and action
                        user_mask = (df_log['name'] == user['name']) & (df_log['action'] == action)
                        if user_mask.any():
                            last_idx = df_log[user_mask].index[-1]
                            df_log.at[last_idx, 'remaining_amount'] = f"₹{final_user['remaining_amount']:,.2f}"
                            df_log.at[last_idx, 'total_interest'] = f"₹{new_interest:,.2f}"
                            df_log.to_excel(LOG_FILE, index=False)
                    
                    # Show that the interest update has been processed
                    st.success(f"Interest update of ₹{abs(amt_signed):,.2f} recorded and recalculated chronologically")
                    st.info(f"Interest: ₹{old_interest:,.2f} → ₹{new_interest:,.2f}")
                    st.rerun()

                elif action == "Interest Rate":
                    old_rate = user['monthly_interest_rate']
                    new_rate = max(old_rate + amt_signed, 0)
                    
                    # Log the rate change first (chronological recalculation depends on log entries)
                    log_action(user, action, amt_signed, action, {
                        "remaining_amount": f"₹{user['remaining_amount']:,.2f}",
                        "total_interest": f"₹{user['total_interest']:,.2f}",
                        "monthly_interest_rate": f"{old_rate}%"
                    }, time=custom_dt.strftime("%Y-%m-%d"))
                    
                    # Update rate in DB first
                    collection.update_one(
                        {"_id": user["_id"]},
                        {"$set": {
                            "monthly_interest_rate": new_rate
                        }}
                    )
                    
                    # Recalculate interest with new rate chronologically
                    user_updated = collection.find_one({"_id": user["_id"]})
                    apply_monthly_interest(user_updated, until_date=datetime.now())
                    
                    # Get final values after recalculation
                    final_user = collection.find_one({"_id": user["_id"]})
                    
                    # Update the log entry with final values
                    if os.path.exists(LOG_FILE):
                        df_log = pd.read_excel(LOG_FILE)
                        # Find the most recent entry for this user and action
                        user_mask = (df_log['name'] == user['name']) & (df_log['action'] == action)
                        if user_mask.any():
                            last_idx = df_log[user_mask].index[-1]
                            df_log.at[last_idx, 'remaining_amount'] = f"₹{final_user['remaining_amount']:,.2f}"
                            df_log.at[last_idx, 'total_interest'] = f"₹{final_user['total_interest']:,.2f}"
                            df_log.at[last_idx, 'monthly_interest_rate'] = f"{new_rate}%"
                            df_log.to_excel(LOG_FILE, index=False)
                    
                    # Show that the rate change has been processed
                    st.success(f"Interest rate change of {amt_signed:+.2f}% recorded and recalculated chronologically")
                    st.info(f"Rate: {old_rate}% → {new_rate}%")
                    st.rerun()




elif selected == "View":
    st.markdown(create_dashboard_header("𝄜 Database View"), unsafe_allow_html=True)
    
    # Clean up orphaned log entries first
    clean_orphaned_log_entries()
    
    user_filter = get_user_filter()
    names = sorted([u['name'] for u in collection.find(user_filter)])
    
    # All users see only their own data
    if not names:
        st.warning("No users found. Please add a user first.")
        st.stop()
    selected = st.selectbox("User", names)
    timeline = []

    # Add loan creation data
    for u in collection.find(user_filter):
        if u['name'] != selected:
            continue
        timeline.append({
            "name": u['name'],
            "action": "Created",
            "amount_changed": None,
            "remaining_amount": f"₹{u['remaining_amount']:,.2f}",
            "total_interest": f"₹{u['total_interest']:,.2f}",
            "loan_start_date": u.get('loan_start_date', '').strftime('%Y-%m-%d') if isinstance(u.get('loan_start_date'), datetime) else u.get('loan_start_date', ''),
            "comment": "Loan created",
            "last_amount_update": u.get('last_interest_update', ''),
            "time": u.get('loan_start_date', '')
        })

    # Add log data from Excel (only for users that exist in database)
    if os.path.exists(LOG_FILE):
        df_logs = pd.read_excel(LOG_FILE)
        df_logs = df_logs[df_logs['name'] == selected]
        
        # Filter out log entries for users that don't exist in database and respect user permissions
        existing_users = {u['name'] for u in collection.find(user_filter)}
        df_logs = df_logs[df_logs['name'].isin(existing_users)]
        
        # All users see only their own data
        if 'created_by_user' in df_logs.columns:
            df_logs = df_logs[df_logs['created_by_user'] == login_state.get("username")]
        else:
            # If created_by_user column doesn't exist, show empty dataframe
            df_logs = df_logs.iloc[0:0]  # Empty dataframe with same structure
        
        for _, row in df_logs.iterrows():
            timeline.append(row.to_dict())

    # Convert to DataFrame
    df = pd.DataFrame(timeline)

    # Ensure 'date' column is clean and shows actual dates
    if 'time' in df.columns:
        df['date'] = pd.to_datetime(df['time'], errors='coerce').dt.strftime("%Y-%m-%d")
    else:
        df['date'] = ""

    # Clean up the data for better display
    for idx, row in df.iterrows():
        action = row.get('action', '')
        
        # For Interest Auto Update entries, clean up the display
        if action == 'Interest Auto Update':
            # Extract the actual date from time column
            time_val = row.get('time', '')
            if time_val:
                try:
                    actual_date = pd.to_datetime(time_val).strftime("%Y-%m-%d")
                    df.at[idx, 'date'] = actual_date
                except:
                    pass
        
        # For Payment entries, combine amount with total interest info
        elif action == 'Payment':
            amount_changed = row.get('amount_changed', '')
            total_interest = row.get('total_interest', '')
            if amount_changed and total_interest and '→' in str(total_interest):
                # Extract final total interest
                final_interest = str(total_interest).split('→')[-1].strip()
                df.at[idx, 'comment'] = f"{row.get('comment', '')} (Total Interest: {final_interest})"
                # Clear total_interest to avoid repetition
                df.at[idx, 'total_interest'] = ""

    # Keep only desired display columns
    display_cols = ['name', 'action', 'amount_changed', 'remaining_amount', 'total_interest', 'comment', 'date']
    df_display = df[[col for col in display_cols if col in df.columns]].copy()

    # Sort by actual date
    df_display['sort_time'] = pd.to_datetime(df_display['date'], errors='coerce')
    df_display = df_display.sort_values(by='sort_time').drop(columns=['sort_time'])

    # ✅ Filter out unwanted entries from display (keep only real business transactions)
    df_display = df_display[~df_display['action'].isin(['Summary After Delete'])]
    
    # ✅ Remove duplicate entries more precisely
    # Keep only one 'Created' entry per user (the first one)
    created_mask = df_display['action'] == 'Created'
    df_created = df_display[created_mask].drop_duplicates(subset=['name'], keep='first')
    df_others = df_display[~created_mask]
    
    # Combine back
    df_display = pd.concat([df_created, df_others], ignore_index=True)
    
    # Sort again by date after combining
    df_display['sort_time'] = pd.to_datetime(df_display['date'], errors='coerce')
    df_display = df_display.sort_values(by='sort_time').drop(columns=['sort_time'])

    # Clean currency fields
    import re
    def extract_final_amount(x):
        try:
            if pd.isna(x):
                return None
            x = str(x)
            if "→" in x:
                x = x.split("→")[-1]
            x = re.sub(r"\(.*?\)", "", x)
            x = x.replace("₹", "").replace(",", "").strip()
            return float(x)
        except:
            return None

    df_display['remaining_amount'] = df_display['remaining_amount'].apply(extract_final_amount)
    df_display['total_interest'] = df_display['total_interest'].apply(extract_final_amount)

    # ✅ Use last non-null values instead of sum, but ignore 'Created' and 'Summary After Delete' actions
    valid_rows = df_display[~df_display['action'].isin(['Created', 'Summary After Delete'])]

    # For the selected user, always show the current state from the database
    user_query = user_filter.copy()
    user_query["name"] = selected
    user_doc = collection.find_one(user_query)
    total_remaining = user_doc["remaining_amount"] if user_doc else 0
    total_interest = user_doc["total_interest"] if user_doc else 0

    # ---------- Inject Correct Total Interest from Summary ----------
    if selected != "All":
        df_interest = df_display[df_display['action'].isin(["Interest Auto Update", "Interest Update"])].copy()
        df_interest['date'] = pd.to_datetime(df_interest['date'], errors='coerce')
        df_interest = df_interest.sort_values(by='date')

        running_total = 0.0

        for _, row in df_interest.iterrows():
            date = row['date'].strftime('%Y-%m-%d') if not pd.isna(row['date']) else ''
            action = row['action']
            amt_raw = str(row.get('amount_changed', '')).replace(",", "").replace("₹", "").strip()
            amt = 0.0

            # Handle both − (unicode) and - (ASCII) negative symbols
            if "−" in amt_raw or "-" in amt_raw:
                amt = -float(''.join(c for c in amt_raw if c.isdigit() or c == '.' or c == '-'))
            else:
                amt = float(''.join(c for c in amt_raw if c.isdigit() or c == '.'))

            if action == "Interest Update":
                running_total -= amt
            else:
                running_total += amt

            total_interest_str = f"₹{running_total:,.2f}"

            # Update df_display's total_interest at this date and action
            match_idx = df_display[
                (df_display['action'] == action) &
                (df_display['date'] == date)
            ].index

            if not match_idx.empty:
                df_display.at[match_idx[0], 'total_interest'] = total_interest_str

    # ---------- Summary View for selected user ----------
    st.markdown(create_section_header(f"📊 Summary for {selected}"), unsafe_allow_html=True)
    df_user = df_display[df_display['name'] == selected].copy()
    df_user = df_user[df_user['remaining_amount'].notnull() & df_user['total_interest'].notnull()]
    if not df_user.empty:
        latest_row = df_user.iloc[-1]
        remaining_amount = latest_row['remaining_amount']
        total_interest = latest_row['total_interest']  # already formatted string
        user_query = user_filter.copy()
        user_query["name"] = selected
        user = collection.find_one(user_query)
        st.markdown(f"""
            - **Amount:** ₹{user.get("principal", 0.0)}
            - **Remaining Amount:** ₹{remaining_amount:,.2f}  
            - **Total Interest:** {total_interest}  
            - **Interest Rate:** {user.get('monthly_interest_rate', 0)}%
        """)
    else:
        st.warning("No data available to summarize.")

    # ------------------- Delete Action Section -------------------
    if selected != "All":
        st.markdown(create_section_header("🗑️ Transaction Management"), unsafe_allow_html=True)
        st.markdown("Remove incorrect transactions:")

        if os.path.exists(LOG_FILE):
            df_logs = pd.read_excel(LOG_FILE)
            df_user_logs = df_logs[df_logs['name'] == selected].copy()
            df_user_logs = df_user_logs[~df_user_logs['action'].isin(['Created', 'Summary After Delete'])]

            if not df_user_logs.empty:
                transaction_options = []
                transaction_indices = []

                for idx, row in df_user_logs.iterrows():
                    action = row['action']
                    amount = row.get('amount_changed', '')
                    date = row.get('time', '') or row.get('timestamp', '') or row.get('date', '')
                    comment = row.get('comment', '')

                    if date and str(date) != 'nan' and str(date).strip() and str(date) != 'None':
                        try:
                            if isinstance(date, str):
                                date_clean = pd.to_datetime(date).strftime('%Y-%m-%d')
                            else:
                                date_clean = pd.to_datetime(date).strftime('%Y-%m-%d')
                        except:
                            date_clean = str(date)[:10]
                    else:
                        date_clean = "No date"

                    if amount and amount != 'nan':
                        amount_clean = str(amount).replace('+₹', '₹').replace('-₹', '-₹')
                    else:
                        amount_clean = ""

                    if comment and comment != 'nan' and str(comment).strip():
                        comment_clean = str(comment).strip()
                        if len(comment_clean) > 30:
                            comment_clean = comment_clean[:27] + "..."
                    else:
                        comment_clean = ""

                    if action == "Interest Auto Update":
                        if "Feb 2025" in comment_clean:
                            display_text = f"Auto Interest - Feb 2025 ({amount_clean})"
                        elif "Mar 2025" in comment_clean:
                            display_text = f"Auto Interest - Mar 2025 ({amount_clean})"
                        elif "Apr 2025" in comment_clean:
                            display_text = f"Auto Interest - Apr 2025 ({amount_clean})"
                        elif "May 2025" in comment_clean:
                            display_text = f"Auto Interest - May 2025 ({amount_clean})"
                        elif "Jun 2025" in comment_clean:
                            display_text = f"Auto Interest - Jun 2025 ({amount_clean})"
                        elif "Jul 2025" in comment_clean:
                            display_text = f"Auto Interest - Jul 2025 ({amount_clean})"
                        else:
                            display_text = f"Auto Interest ({amount_clean})"
                    elif action == "Payment":
                        display_text = f"Payment {amount_clean} - {date_clean}"
                    elif action == "Interest Update":
                        display_text = f"Interest Update {amount_clean} - {date_clean}"
                    elif action == "Interest Rate":
                        display_text = f"Rate Change {amount_clean} - {date_clean}"
                    else:
                        display_text = f"{action} {amount_clean} - {date_clean}"

                    if comment_clean and comment_clean not in ["Payment", "Interest Update", "Interest Rate"]:
                        display_text += f" ({comment_clean})"

                    transaction_options.append(display_text)
                    transaction_indices.append(idx)

                col1, col2 = st.columns([4, 1])
                with col1:
                    selected_transaction = st.selectbox(
                        "Select transaction to remove:",
                        options=range(len(transaction_options)),
                        format_func=lambda x: transaction_options[x],
                        key="transaction_selector"
                    )
                with col2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("Remove", key="remove_transaction"):
                        selected_idx = transaction_indices[selected_transaction]
                        if delete_action(selected_idx, selected):
                            st.success("Transaction removed successfully!")
                            st.rerun()
                        else:
                            st.error("Failed to remove transaction!")
            else:
                st.info("No transactions found for this user.")

    # ------------------- Summary Row -------------------
    summary_row = {
        "name": "TOTAL",
        "action": "—",
        "amount_changed": "—",
        "remaining_amount": f"₹{total_remaining:,.2f}",
        "total_interest": total_interest,  # already a formatted string
        "comment": "Current totals from database",
        "date": "—"
    }

    df_display = pd.concat([df_display, pd.DataFrame([summary_row])], ignore_index=True)

    # ------------------- Display -------------------
    st.dataframe(df_display, use_container_width=True)


    # ------------------- Excel Export -------------------
    def export_to_excel(dataframe):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            dataframe.to_excel(writer, index=False, sheet_name='CombinedData')

            from openpyxl.styles import Font, Alignment, PatternFill
            import openpyxl

            workbook = writer.book
            worksheet = writer.sheets['CombinedData']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            for col_num, column_title in enumerate(dataframe.columns, 1):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                cell = worksheet[f'{col_letter}1']
                cell.font = header_font
                cell.fill = header_fill
                worksheet.column_dimensions[col_letter].width = max(15, len(column_title) + 3)

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        return output.getvalue()

    excel_bytes = export_to_excel(df_display)
    st.download_button(
        label="📥 Download Data as Excel",
        data=excel_bytes,
        file_name="Data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
# Data Entry Page
elif selected == "Data Entry":
    st.markdown(create_dashboard_header("👨‍💻 Data Entry"), unsafe_allow_html=True)
    
    st.subheader("Manual Data Entry")
    
    DATA_ENTRY_FILE = "data_entry_records.xlsx"

    with st.form("data_entry_form", clear_on_submit=True):
        entry_date = st.date_input("Date", value=datetime.today())
        name = st.text_input("Name")
        amount = st.number_input("Amount", min_value=0.0, format="%.2f")
        interest = st.number_input("Interest", min_value=0.0, format="%.2f")
        total = st.number_input("Total", min_value=0.0, format="%.2f")
        submit_btn = st.form_submit_button("Submit")

        if submit_btn:
            if not name.strip():
                st.error("Name is required!")
            else:
                # Create data entry document
                entry_doc = {
                    "date": datetime.combine(entry_date, datetime.min.time()),  # Convert date to datetime
                    "name": name.strip(),
                    "amount": float(amount),
                    "interest": float(interest),
                    "total": float(total),
                    "created_by_user": login_state.get("username"),
                    "created_at": datetime.now()
                }
                
                # Save to MongoDB
                data_entries_col.insert_one(entry_doc)
                
                # Also save to Excel for backup
                new_data = pd.DataFrame([{
                    "Date": entry_date.strftime("%Y-%m-%d"),
                    "Name": name.strip(),
                    "Amount": amount,
                    "Interest": interest,
                    "Total": total,
                    "created_by_user": login_state.get("username")
                }])

                if os.path.exists(DATA_ENTRY_FILE):
                    df_existing = pd.read_excel(DATA_ENTRY_FILE)
                    df_combined = pd.concat([df_existing, new_data], ignore_index=True)
                else:
                    df_combined = new_data

                df_combined.to_excel(DATA_ENTRY_FILE, index=False)
                st.success("✅ Data saved to database and Excel!")
                st.rerun()

    # New feature: Filter by user - Read from Database
    st.subheader("🔍 View User Entries")
    
    # Get data from MongoDB
    user_filter = {"created_by_user": login_state.get("username")}
    entries = list(data_entries_col.find(user_filter).sort("date", -1))
    
    if not entries:
        st.info("No entries found. Add some entries using the form above.")
        st.stop()
    
    # Convert to DataFrame
    df_entries = pd.DataFrame(entries)
    df_entries['Date'] = pd.to_datetime(df_entries['date']).dt.strftime('%Y-%m-%d')
    df_entries['Name'] = df_entries['name']
    df_entries['Amount'] = df_entries['amount']
    df_entries['Interest'] = df_entries['interest']
    df_entries['Total'] = df_entries['total']
    
    # Display columns in proper order
    display_df = df_entries[['Date', 'Name', 'Amount', 'Interest', 'Total']].copy()
    
    # Get unique names (filter out NaN values)
    unique_names = sorted([name for name in df_entries['name'].unique() if pd.notna(name)])
    
    selected_name = st.selectbox("Select a user to view their entries", unique_names)
    
    user_df = display_df[display_df['Name'] == selected_name]
    st.write(f"### All entries for {selected_name}")
    
    # Calculate and display summary stats
    total_amount = user_df['Amount'].sum()
    total_interest = user_df['Interest'].sum()
    overall_total = user_df['Total'].sum()
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Amount", f"₹{total_amount:,.2f}")
    with col2:
        st.metric("Total Interest", f"₹{total_interest:,.2f}")
    with col3:
        st.metric("Overall Total", f"₹{overall_total:,.2f}")
    
    st.dataframe(user_df, use_container_width=True)
    
    st.subheader("✏️ Edit Data")
    
    # Get data for editing from MongoDB
    if entries:
        # Convert MongoDB data to editable DataFrame
        edit_df = display_df.copy()
        edit_df['_id'] = [str(entry['_id']) for entry in entries]  # Keep MongoDB _id for updates
        edit_df['Delete'] = False
        
        edited_df = st.data_editor(
            edit_df,
            use_container_width=True,
            key="edit_data_entry",
            column_config={
                "Delete": st.column_config.CheckboxColumn("Delete?"),
                "_id": None  # Hide the _id column
            }
        )
        
        # Check if any rows are marked for deletion
        if edited_df['Delete'].any():
            if st.button("Confirm Delete Selected Rows"):
                # Get rows marked for deletion
                rows_to_delete = edited_df[edited_df['Delete']]
                
                # Delete from MongoDB
                for _, row in rows_to_delete.iterrows():
                    from bson import ObjectId
                    data_entries_col.delete_one({"_id": ObjectId(row['_id'])})
                
                # Update Excel backup
                if os.path.exists(DATA_ENTRY_FILE):
                    df_excel = pd.read_excel(DATA_ENTRY_FILE)
                    if 'created_by_user' in df_excel.columns:
                        # Remove corresponding entries from Excel
                        for _, row in rows_to_delete.iterrows():
                            mask = ((df_excel['Date'] == row['Date']) & 
                                   (df_excel['Name'] == row['Name']) & 
                                   (df_excel['Amount'] == row['Amount']) &
                                   (df_excel['created_by_user'] == login_state.get("username")))
                            df_excel = df_excel[~mask]
                        df_excel.to_excel(DATA_ENTRY_FILE, index=False)
                    
                st.success("✅ Selected rows deleted from database!")
                st.rerun()
        
        # Check for other edits (non-deletion changes)
        elif not edited_df.drop(columns=['Delete', '_id']).equals(edit_df.drop(columns=['Delete', '_id'])):
            # Update MongoDB with changes
            changes_made = False
            for idx, row in edited_df.iterrows():
                if not row['Delete']:  # Skip deleted rows
                    original_row = edit_df.iloc[idx]
                    # Check if this row was actually changed
                    if not (row.drop(['Delete', '_id']).equals(original_row.drop(['Delete', '_id']))):
                        from bson import ObjectId
                        # Update in MongoDB
                        data_entries_col.update_one(
                            {"_id": ObjectId(row['_id'])},
                            {"$set": {
                                "date": pd.to_datetime(row['Date']),  # Keep as datetime, not date
                                "name": row['Name'],
                                "amount": float(row['Amount']),
                                "interest": float(row['Interest']),
                                "total": float(row['Total'])
                            }}
                        )
                        changes_made = True
            
            if changes_made:
                # Update Excel backup
                if os.path.exists(DATA_ENTRY_FILE):
                    # Recreate Excel from MongoDB data
                    all_entries = list(data_entries_col.find({"created_by_user": login_state.get("username")}))
                    if all_entries:
                        df_backup = pd.DataFrame([{
                            "Date": entry['date'].strftime("%Y-%m-%d") if hasattr(entry['date'], 'strftime') else str(entry['date']),
                            "Name": entry['name'],
                            "Amount": entry['amount'],
                            "Interest": entry['interest'],
                            "Total": entry['total'],
                            "created_by_user": entry['created_by_user']
                        } for entry in all_entries])
                        
                        # Read existing Excel and update user's portion
                        if os.path.exists(DATA_ENTRY_FILE):
                            df_excel = pd.read_excel(DATA_ENTRY_FILE)
                            if 'created_by_user' in df_excel.columns:
                                # Remove user's old data
                                df_excel = df_excel[df_excel['created_by_user'] != login_state.get("username")]
                                # Add user's updated data
                                df_excel = pd.concat([df_excel, df_backup], ignore_index=True)
                            else:
                                df_excel = df_backup
                        else:
                            df_excel = df_backup
                        
                        df_excel.to_excel(DATA_ENTRY_FILE, index=False)
                
                st.success("✅ Changes saved to database!")
                st.rerun()
    else:
        st.info("No entries to edit. Add some entries first.")
    
    # Add Excel download button with styling at the bottom
    st.markdown("---")  # Add a horizontal line for separation
    
    def export_to_excel(dataframe):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            dataframe.to_excel(writer, index=False, sheet_name='CombinedData')

            from openpyxl.styles import Font, Alignment, PatternFill
            import openpyxl

            workbook = writer.book
            worksheet = writer.sheets['CombinedData']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            for col_num, column_title in enumerate(dataframe.columns, 1):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                cell = worksheet[f'{col_letter}1']
                cell.font = header_font
                cell.fill = header_fill
                worksheet.column_dimensions[col_letter].width = max(15, len(column_title) + 3)

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        return output.getvalue()

    # Export data from MongoDB
    user_filter = {"created_by_user": login_state.get("username")}
    export_entries = list(data_entries_col.find(user_filter).sort("date", -1))
    
    if export_entries:
        export_df = pd.DataFrame([{
            "Date": entry['date'].strftime("%Y-%m-%d") if hasattr(entry['date'], 'strftime') else str(entry['date']),
            "Name": entry['name'],
            "Amount": entry['amount'],
            "Interest": entry['interest'],
            "Total": entry['total']
        } for entry in export_entries])
    else:
        # Empty dataframe if no entries
        export_df = pd.DataFrame(columns=["Date", "Name", "Amount", "Interest", "Total"])
    
    excel_bytes = export_to_excel(export_df)
    st.download_button(
        label="📥 Download Data as Excel",
        data=excel_bytes,
        file_name="Data_entry.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("No records yet. Use the form above to add entries.")
        
client.close()
